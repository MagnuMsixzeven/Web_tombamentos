import openpyxl, json, re
from datetime import datetime
from collections import Counter

wb = openpyxl.load_workbook(r'C:\Users\Usuario\Desktop\Patrimonios 2023 e 2025.xlsx')

tombamentos = []
historico = []
next_id = 1
marcas_set = set()
produtos_set = set()

def normalizar_setor(setor):
    if not setor: return 'TI'
    s = str(setor).strip()
    s_upper = s.upper()
    if 'TI' in s_upper and ('RAFAEL' in s_upper or 'ASTIR' in s_upper or s_upper == 'TI'):
        return 'TI'
    if 'GUIA' in s_upper: return 'Guias'
    if 'DIREX' in s_upper or 'DIRETORIA EXECUTIVA' in s_upper: return 'DIREX'
    if 'FATURAMENTO' in s_upper or 'FATUR' in s_upper: return 'Faturamento'
    if 'CADASTRO' in s_upper: return 'Cadastro'
    if 'AMBUL' in s_upper: return 'Ambulat\u00f3rio'
    if 'FINANC' in s_upper: return 'Financeiro'
    if 'COMPRA' in s_upper: return 'Compras'
    if 'JUR' in s_upper: return 'Jur\u00eddico'
    if s_upper in ('RH',) or 'RECURSO' in s_upper: return 'RH'
    if 'ALMOX' in s_upper: return 'Almoxarifado'
    if 'RECEP' in s_upper: return 'Recep\u00e7\u00e3o'
    if 'PRESID' in s_upper: return 'Presid\u00eancia'
    if 'CONT' in s_upper and 'ABIL' in s_upper: return 'Contabilidade'
    if 'ODONTO' in s_upper: return 'Odontologia'
    if 'FISIO' in s_upper: return 'Fisioterapia'
    if 'FARM' in s_upper: return 'Farm\u00e1cia'
    if 'LABORA' in s_upper: return 'Laborat\u00f3rio'
    if 'RAIO' in s_upper or 'RX' in s_upper or 'RADIO' in s_upper: return 'Raio-X'
    if 'ENFERM' in s_upper: return 'Enfermagem'
    if 'PSICO' in s_upper: return 'Psicologia'
    if 'NUTRI' in s_upper: return 'Nutri\u00e7\u00e3o'
    if 'AUDIO' in s_upper: return 'Audiologia'
    if 'OTORRINO' in s_upper: return 'Otorrinolaringologia'
    if 'COMUNIC' in s_upper: return 'Comunica\u00e7\u00e3o'
    if 'SALA' in s_upper and 'REUNI' in s_upper: return 'Sala de Reuni\u00e3o'
    if 'COZINHA' in s_upper: return 'Cozinha'
    if 'PORTARIA' in s_upper: return 'Portaria'
    if 'AGEND' in s_upper: return 'Agendamento'
    if 'ACOLH' in s_upper: return 'Acolhimento'
    if 'SERV' in s_upper and 'SOCIAL' in s_upper: return 'Servi\u00e7o Social'
    if 'ASSES' in s_upper: return 'Assessoria'
    if 'MEDIC' in s_upper or 'CONSULT' in s_upper: return 'Consult\u00f3rio M\u00e9dico'
    return s.title()

def normalizar_material(mat):
    if not mat: return 'Outros'
    m = str(mat).strip().upper()
    mapping = {
        'COMPUTADOR': 'CPU', 'PC': 'CPU', 'DESKTOP': 'CPU', 'GABINETE': 'CPU',
        'MONITOR': 'Monitor', 'TELA': 'Monitor',
        'NOBREAK': 'Nobreak', 'NO BREAK': 'Nobreak', 'NO-BREAK': 'Nobreak',
        'SWITCH': 'Switch',
        'IMPRESSORA': 'Impressora',
        'ROTEADOR': 'Roteador', 'ROUTER': 'Roteador',
        'DVR': 'DVR',
        'CAMERA': 'C\u00e2mera', 'C\u00c2MERA': 'C\u00e2mera',
        'RACK': 'Rack',
        'MICROFONE': 'Microfone',
        'TECLADO': 'Teclado',
        'MOUSE': 'Mouse',
        'TELEFONE': 'Telefone', 'APARELHO TELEFONICO': 'Telefone',
        'CELULAR': 'Celular',
        'NOTEBOOK': 'Notebook',
        'ESTABILIZADOR': 'Estabilizador',
        'CAIXA DE SOM': 'Caixa de Som',
        'HD EXTERNO': 'HD Externo',
        'PROJETOR': 'Projetor', 'DATA SHOW': 'Projetor',
        'ACCESS POINT': 'Access Point',
        'TELEVISOR': 'TV', 'TV': 'TV',
    }
    for key, val in mapping.items():
        if key in m:
            return val
    return str(mat).strip().title()

def normalizar_status(status):
    if not status: return 'Ativo'
    s = str(status).strip().upper()
    if 'BAIXA' in s or 'DESCART' in s or 'PERDIDO' in s or 'MAU' in s:
        return 'Em Baixa'
    return 'Ativo'

def parse_data(d):
    if not d: return datetime.now().isoformat()
    if isinstance(d, datetime):
        return d.isoformat()
    try:
        return datetime.strptime(str(d).strip(), '%Y-%m-%d').isoformat()
    except:
        try:
            return datetime.strptime(str(d).strip(), '%d/%m/%Y').isoformat()
        except:
            return datetime.now().isoformat()

def extrair_tombamento_num(pat):
    if not pat: return 0
    s = str(pat).strip()
    match = re.match(r'^\d{2}/(\d+)$', s)
    if match:
        return int(match.group(1))
    nums = re.findall(r'(\d+)', s)
    if nums:
        return int(nums[0])
    return 0

# Sheet 1: NOVO PATRIMONIO
ws1 = wb['NOVO PATRIMONIO']
for row in ws1.iter_rows(min_row=2, values_only=True):
    pat = row[0]
    if not pat: continue
    data_cad = row[1]
    material = row[2]
    cor = row[3]
    desc = row[4]
    marca_raw = row[5]
    modelo = row[6]
    setor_raw = row[7]
    nserie = row[8]
    baixa = row[9]

    produto = normalizar_material(material)
    marca = str(marca_raw).strip().upper() if marca_raw else 'N/I'
    setor = normalizar_setor(setor_raw)
    status = normalizar_status(baixa)
    serie = str(nserie).strip() if nserie else 'S/N'
    num_tomb = extrair_tombamento_num(pat)
    data_iso = parse_data(data_cad)

    detalhes = ''
    parts = []
    if desc: parts.append(str(desc).strip())
    if cor: parts.append('Cor: ' + str(cor).strip())
    if modelo: parts.append('Modelo: ' + str(modelo).strip())
    detalhes = ' | '.join(parts)

    nome = str(pat).strip()

    tombamentos.append({
        'id': next_id,
        'numero_tombamento': num_tomb,
        'produto': produto,
        'marca': marca,
        'numero_serie': serie,
        'status': status,
        'setor': setor,
        'data_cadastro': data_iso,
        'nome': nome,
        'detalhes': detalhes
    })

    marcas_set.add(marca)
    produtos_set.add(produto)

    historico.append({
        'tombamento': num_tomb,
        'tipo': 'Importa\u00e7\u00e3o',
        'de': '\u2014',
        'para': setor,
        'por': 'TI',
        'data': data_iso
    })

    next_id += 1

novo_count = next_id - 1
print('NOVO PATRIMONIO: {} registros'.format(novo_count))
novo_nums = sorted(set(t['numero_tombamento'] for t in tombamentos))
print('  Tombamentos: {} a {}'.format(novo_nums[0], novo_nums[-1]))

# Sheet 2: ANTIGO PATRIMONIO
ws2 = wb['ANTIGO PATRIMONIO']
antigo_start = next_id
for row in ws2.iter_rows(min_row=2, values_only=True):
    numpat = row[0]
    if not numpat: continue
    data_cad = row[1]
    situacao = row[2]
    material = row[3]
    cor = row[4]
    desc = row[5]
    marca_raw = row[6]
    modelo = row[7]
    setor_raw = row[8]
    nserie = row[10] if len(row) > 10 else None

    produto = normalizar_material(material)
    marca = str(marca_raw).strip().upper() if marca_raw else 'N/I'
    setor = normalizar_setor(setor_raw)
    status = normalizar_status(situacao)
    serie = str(nserie).strip() if nserie else 'S/N'
    num_tomb = extrair_tombamento_num(numpat)
    data_iso = parse_data(data_cad)

    detalhes = ''
    parts = []
    if desc: parts.append(str(desc).strip())
    if cor: parts.append('Cor: ' + str(cor).strip())
    if modelo: parts.append('Modelo: ' + str(modelo).strip())
    if situacao: parts.append('Estado: ' + str(situacao).strip())
    detalhes = ' | '.join(parts)

    nome = 'PAT-' + str(numpat).strip()

    tombamentos.append({
        'id': next_id,
        'numero_tombamento': num_tomb,
        'produto': produto,
        'marca': marca,
        'numero_serie': serie,
        'status': status,
        'setor': setor,
        'data_cadastro': data_iso,
        'nome': nome,
        'detalhes': detalhes
    })

    marcas_set.add(marca)
    produtos_set.add(produto)

    historico.append({
        'tombamento': num_tomb,
        'tipo': 'Importa\u00e7\u00e3o',
        'de': '\u2014',
        'para': setor,
        'por': 'TI',
        'data': data_iso
    })

    next_id += 1

antigo_count = next_id - antigo_start
print('ANTIGO PATRIMONIO: {} registros'.format(antigo_count))

all_nums = sorted(set(t['numero_tombamento'] for t in tombamentos))
print('')
print('Total: {} registros'.format(len(tombamentos)))
print('Range tombamentos: {} a {}'.format(all_nums[0], all_nums[-1]))
print('Tombamentos unicos: {}'.format(len(all_nums)))

counter = Counter(t['numero_tombamento'] for t in tombamentos)
dups = {k: v for k, v in counter.items() if v > 1}
if dups:
    print('')
    print('Duplicados ({}):'.format(len(dups)))
    for k in sorted(dups):
        items = [t for t in tombamentos if t['numero_tombamento'] == k]
        info = [(t['nome'], t['produto']) for t in items]
        print('  #{}: {}'.format(k, info))

output = {
    'tombamentos': tombamentos,
    'historico': historico,
    'marcas': sorted(list(marcas_set)),
    'produtos': sorted(list(produtos_set))
}

with open(r'c:\Users\Usuario\Desktop\portifolio\_Web_Tombamento\public\dados_importacao.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print('')
print('Arquivo dados_importacao.json atualizado!')
