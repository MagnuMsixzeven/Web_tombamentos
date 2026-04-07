import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Usuario\Desktop\Patrimonios 2023 e 2025.xlsx')

ws1 = wb['NOVO PATRIMONIO']
print('=== NOVO PATRIMONIO - coluna PAT (primeiros 20) ===')
pats = []
for row in ws1.iter_rows(min_row=2, max_row=400, values_only=True):
    if row[0]:
        pats.append(str(row[0]).strip())
        if len(pats) <= 20:
            print(f'  {row[0]}')

print(f'\nTotal: {len(pats)}')
if pats:
    print(f'Ultimo: {pats[-1]}')

prefixos = set()
for p in pats:
    parts = p.split('/')
    if len(parts) >= 2:
        prefixos.add(parts[0])
print(f'Prefixos de ano: {sorted(prefixos)}')

ws2 = wb['ANTIGO PATRIMONIO']
print('\n=== ANTIGO PATRIMONIO - coluna No do Patrimonio (primeiros 20) ===')
pats2 = []
for row in ws2.iter_rows(min_row=2, max_row=100, values_only=True):
    if row[0]:
        pats2.append(str(row[0]).strip())
        if len(pats2) <= 20:
            print(f'  {row[0]}')

print(f'\nTotal: {len(pats2)}')
if pats2:
    print(f'Ultimo: {pats2[-1]}')
